from base64 import b64encode
from datetime import date
from delta.tables import DeltaTable as Δ
from functools import reduce, partial
import json 
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import pandas as pd
from pandas import DataFrame as pd_DF
from pathlib import WindowsPath, Path
from pyspark.sql import (functions as F, types as T, Window as W, 
    Row, Column, DataFrame as spk_DF) 
import re
from requests import Request
import sys 
from typing import Union
   

def dict_get2(a_dict: dict, ls_keys:list): 
    b_dict = a_dict.copy()
    if ls_keys[-1] not in a_dict: 
        b_dict[None] = ls_keys[-1]

    ls_keys_0 = [k for k in ls_keys if k in a_dict]
    a_key = ls_keys_0[0] if ls_keys_0 else None
    a_val = a_dict.get(a_key)
    return a_val
            

def reload_from_asterisk(module:str):
    a_mod = reload(sys.modules[module])
    vars().update(a_mod.__dict__)
    print(f'Module {module} reloaded.')


def str_camel_to_snake(cameled:str):
    subbed = re.sub('(.)([A-Z][a-z]+)',  r'\1_\2', cameled)
    snaked = re.sub('([a-z0-9])([A-Z])', r'\1_\2', subbed).lower()
    return snaked


def str_snake_to_camel(snaked:str, first_word_too=False):
    first, *others = snaked.split('_')
    first_too = first.title() if first_word_too else first.lower()
    cameled = first_too + ''.join(word.title() for word in others)
    return cameled


def snake_2_camel(snake_str:str):
    first, *others = snake_str.split('_')
    return ''.join([first.lower(), *map(str.title, others)])
      

def select_subdict(a_dict:str, sub_keys:list):
    if not sub_keys.is_subset(a_dict.keys()):
        raise ValueError('Trying to select a subdict with keys not contained on large dict.')
    sub_dict = dict((k, a_dict[k]) for k in sub_keys)
    return sub_dict


def camelize_dict(snake_dict:dict):
    case_type = {
        dict: camelize_dict,    # Recursive. 
        list: lambda v: list(map(camelize_dict, v)), 
        None: lambda v: v}      # else
    pass_value = MatchCase(case_type, via=type)

    new_dict = dict((k, pass_value(v)) 
            for (k, v) in snake_dict.items())
    return new_dict


def read_excel_table(a_file, sheet:str, table:str=None, **kwargs): 
    if table is None: 
        table = sheet.lower().replace(' ', '_')
    try:
        a_wb = load_workbook(a_file, data_only=True)
    except InvalidFileException: 
        a_wb = load_workbook(shortcut_target(a_file), data_only=True)
    
    a_ws  = a_wb[sheet]

    if table not in a_ws.tables: 
        return None

    a_tbl = a_ws.tables[table]    
    rows_ls = [[ cell.value for cell in row ] for row in a_ws[a_tbl.ref]]
    tbl_df  = pd.DataFrame(data=rows_ls[1:], index=None, columns=rows_ls[0], **kwargs)
    return tbl_df


def shortcut_target(a_file, file_ext:str=None):
    if file_ext is None: 
        if isinstance(a_file, WindowsPath): 
            file_ext = re.findall(r"\.([A-Za-z]{3,4})\.lnk", a_file.name)[0]
        else: 
            raise "Couldn't determine file extension."

    file_regex = fr'(C:\\.*\.{file_ext})'

    with open(a_file, 'r', encoding='ISO-8859-1') as _f: 
        a_path = re.findall(file_regex, _f.read(), flags=re.DOTALL)

    if len(a_path) != 1: 
        raise 'Not unique or no shortcut targets found in link.'
    return a_path[0]


def dict_minus(a_dict:str, key_ls:list, copy=True): 
    b_dict = a_dict.copy() if copy else a_dict
    for key in key_ls: 
        b_dict.pop(key, None)
    return b_dict


encode64 = (lambda a_str: 
    b64encode(a_str.encode('ascii')).decode('ascii'))


def compose(*fs, reverse=True): 
    # Reverse es el estándar:  (f∘g)(x) = f(g(x))
    # Entonces la última de la lista es la primera en aplicarse. 
    if reverse: 
        fs.reverse()

    Φ_0 = lambda xx: xx
    Φ_2 = lambda f, g: (lambda x: f(g(x)))
    return reduce(Φ_2, fs, Φ_0)


def is_na(val: Union[str, float]): 
    if isinstance(val, str): 
        is_it = (val is None)
    elif isinstance(val, float): 
        is_it = np.isnan(val)
    return is_it


def datecols_to_str(a_df:pd_DF):
    b_df = (a_df.select_dtypes(['datetime'])
        .apply(lambda col: col.dt.strftime('%Y-%m-%d'), axis=1))
    return a_df.assign(**b_df)


def dataframe_to_list(a_df:pd_DF, cols_df=None, manual=True): 
    # COLS_DF tiene columnas:
    #  TYPE['datetime', 'date', ...]
    #  ATRIBUTO
    
    if not manual: 
        dt_df = (a_df.select_dtypes(['datetime'])
            .apply(lambda col: col.dt.strftime('%d-%m-%Y'), axis=1))
        the_list = json.loads(a_df.assign(**dt_df).to_json(orient='records'))

    else: 
        the_types = {'bool': 'logical', 'object': 'character'}

        if cols_df is None:
            cols_df = (a_df.dtypes.replace(the_types)
                .to_frame('dtipo').reset_index()
                .rename(columns={'index': 'nombre'})
                .assign(dtype=lambda df: df.dtype))

        ts_cols     = cols_df.query("dtipo == 'datetime'")['nombre']
        date_cols_0 = cols_df.query("dtipo == 'date'")['nombre']

        str_assign  = {col: str 
                for col in ts_cols if col in a_df.columns.values }

        date_assign = {col: a_df[col].dt.strftime('%Y-%m-%d')
                for col in date_cols_0 if col in a_df.columns.values}
        
        b_df = a_df.astype(str_assign).assign(**date_assign)

        the_list = b_df.to_dict(orient='records')
    return the_list


def set_dataframe_types(a_df:pd_DF, cols_df:pd_DF): 
    # ['char', 'numeric', 'date', 'datetime', 'object']
    dtypes = {'char':str, 'numeric':float, 'object':str, 
        'datetime': 'datetime64[ns]', 'date': 'datetime64[ns]', 'delta': str}

    dtyped = { row.database: dtypes[row.dtipo] for row in cols_df.itertuples()
            if row.database in a_df.columns and row.dtipo in dtypes}
    df_typed = a_df.astype(dtyped, errors='ignore')
    return df_typed


def curlify(resp_request:Request): 
    the_items = { 
        'method' : resp_request.method, 
        'headers': ''.join(f"-H '{k}:{v}' " for (k, v) in resp_request.headers.items()), 
        'data'   : resp_request.body, 
        'uri'    : resp_request.url}
    return "curl - X {method} {headers}-d '{data}' '{uri}'".format(**the_items)


class partial2(partial): 
    """
    An improved version of partial which accepts Ellipsis (...) as a placeholder.
    """
    def __call__(self, *args, **keywords):
        keywords = {**self.keywords, **keywords}
        iargs = iter(args)
        args = (next(iargs) if arg is ... else arg for arg in self.args)
        return self.func(*args, *iargs, **keywords)


class MatchCase():
    '''This is the well Known MATCH CASE flow, with functional variant.  
    Use None key, for ELSE statements. 
    switcher = MatchCase(a_dict, via=ff)
    switcher(xx) ~ a_dict[ff(xx)](xx)

    switch ff(xx): 
        case kk:
            vv(xx)
        else: # (None) 
            ww(xx)
    '''
    def __init__(self, case_dict, via=None): 
        self._via = via
        self._case_dict = case_dict.copy()
        
        self.via = via or (lambda xx:xx)
        case_dict.setdefault(None, None)
        self.case_dict = {kk: as_callable(vv) 
            for kk, vv in case_dict.items()}

    def __call__(self, xx): 
        which_k = via(xx)
        which_v = (self.case_dict.get(which_k, case_dict[None]))
        return which_v(xx)
    

def ɸ_constant(cc): 
    return (lambda xx: cc)

def as_callable(yy): 
    ɸɸ = yy if callable(yy) else ɸ_constant(yy)
    return ɸɸ



